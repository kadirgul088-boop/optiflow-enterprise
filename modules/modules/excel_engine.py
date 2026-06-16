from openpyxl import Workbook
from openpyxl.styles import Font

def create_excel_report(
    company_name,
    score,
    maturity,
    company_metrics,
    benchmark_result,
    financial_result
):

    file_name = f"OptiFlow_{company_name}.xlsx"

    wb = Workbook()

    # DASHBOARD
    ws = wb.active
    ws.title = "Dashboard"

    ws["A1"] = "OptiFlow Enterprise Dashboard"
    ws["A1"].font = Font(bold=True, size=14)

    dashboard = [
        ["Metric", "Value"],
        ["OptiFlow Score", score],
        ["Maturity", maturity["Seviye"]],
        ["Wait Rate", company_metrics["wait_rate"]],
        ["OEE", company_metrics["oee"]],
        ["Defect Rate", company_metrics["defect_rate"]],
        ["Line Balance Loss", company_metrics["line_balance_loss"]],
    ]

    for row in dashboard:
        ws.append(row)

    # BENCHMARK
    ws2 = wb.create_sheet("Benchmark")

    for key, value in benchmark_result.items():
        ws2.append([str(key), str(value)])

    # FINANCIAL
    ws3 = wb.create_sheet("Financial")

    for key, value in financial_result.items():
        ws3.append([str(key), str(value)])

    wb.save(file_name)

    return file_name